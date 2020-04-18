import csv
from os import listdir
from cdr_utils import parse_call_forwarding, parse_change_of_software, parse_feature, parse_half_call, parse_intermediate, parse_long_call, parse_qos, parse_queue, parse_standard
from openpyxl import load_workbook, Workbook

path = 'cdrs'
files = listdir(path)

queue_report = Workbook()
hoja = queue_report.active
hoja['A1'] = 'ID'  # CDR field 5
hoja['B1'] = 'Hora'  # CDR field 3
hoja['C1'] = 'Identificador de cola'  # CDR field 7
hoja['D1'] = 'Hora de ingreso a la cola'  # CDR field 8
hoja['E1'] = 'Hora de salida de la cola'  # CDR field 9
hoja['F1'] = 'Abandono'  # CDR field 11
hoja['G1'] = 'Agente de destino'  # CDR field 12
queue_report.save('reports/queue_report.xlsx')

standard_report = Workbook()
sheet = standard_report.active
sheet['A1'] = 'Switch ID'  # CDR field 5
sheet['B1'] = 'ID'  # CDR field 6
sheet['C1'] = 'Hora de inicio'  # CDR field 3
sheet['D1'] = 'Duracion de llamada'  # CDR field 4
sheet['E1'] = 'Numero de origen'  # CDR field 12
sheet['F1'] = 'Numero de destino'  # CDR field 11
sheet['G1'] = 'Resultado'  # CDR field 18
sheet['H1'] = 'Origen en OSV'  # CDR field 40
sheet['I1'] = 'Destino en OSV'  # CDR field 41
sheet['J1'] = 'Hora de atencion'  # CDR field 48
sheet['K1'] = 'Hora de corte'  # CDR field 49
sheet['L1'] = 'Hora de atencion - incoming leg'  # CDR field 50
sheet['M1'] = 'Hora de corte - incoming leg'  # CDR field 51
sheet['N1'] = 'Hora de atencion - outgoing leg'  # CDR field 52
sheet['O1'] = 'Hora de corte - outgoing leg'  # CDR field 53
standard_report.save('reports/standard_report.xlsx')

cdr_types = {
    '00000000': {
        'method': parse_standard,
        'report': standard_report,
        'current_report_line': 2,
    },
    '00000001': {
        'method': parse_intermediate,
        'report': '',
        'current_report_line': 2,
    },
    '00000004': {
        'method': parse_queue,
        'report': queue_report,
        'current_report_line': 2,
    },
    '00000005': {
        'method': parse_call_forwarding,
        'report': '',
        'current_report_line': 2,
    },
    '10000001': {
        'method': parse_long_call,
        'report': '',
        'current_report_line': 2,
    },
    '10000010': {
        'method': parse_change_of_software,
        'report': '',
        'current_report_line': 2,
    },
    '10000100': {
        'method': parse_feature,
        'report': '',
        'current_report_line': 2,
    },
    '10000101': {
        'method': parse_half_call,
        'report': '',
        'current_report_line': 2,
    },
    '10000111': {
        'method': parse_qos,
        'report': '',
        'current_report_line': 2,
    }
}

for file in files:
    start_file = False
    end_file = False
    with open(path + '/' + file, 'r') as csv_file:
        cdr_file = csv.reader(csv_file, delimiter=',')
        for row in cdr_file:
            if len(row) == 0:
                continue
            if row[0][0:7] == 'CREATE:':
                start_file = True
                continue
            if row[0][0:6] == 'CLOSE:':
                end_file = True
                break
            if start_file and not end_file:
                cdr_types[row[1]]['method'](row, cdr_types[row[1]]['report'], cdr_types[row[1]]['current_report_line'])
                cdr_types[row[1]]['current_report_line'] += 1

queue_report.save('reports/queue_report.xlsx')
standard_report.save('reports/standard_report.xlsx')